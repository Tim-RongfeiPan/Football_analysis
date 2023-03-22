#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   get_eventdata.py
@Time    :   2023/03/14 15:07:49
@Author  :   Rongfei Pan
@Version :   1.0
@Contact :   rongfei@kth.se 1838863836prf@gmail.com
@Desc    :   function library for event data
'''

# here put the import lib

from matplotlib.pyplot import step
from openpyxl import *
import datetime
import cv2


class Get_Eventdata:

    def __init__(self, path):
        """initialized by path

        Args:
            path (string): path of xlsx file
        """
        super(Get_Eventdata, self).__init__()
        self.path = path
        self.wb = load_workbook(path)
        self.data = self.wb['Sheet1']

    def get_eventdata_byevent(self, name):
        """get event data by event

        Args:
            name (string): type of event

        Returns:
            list[list]: list of info in row with corresponding name
        """
        name_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 1).value == name:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                name_list.append(name_info)
        return name_list

    def get_eventdata_bytime(self, dtime):
        """get event data by time

        Args:
            dtime (datetime.time): time, Exp: datetime.time(xx, xx, xx)

        Returns:
            list[list]: list with two sublist with closest time. Sublist: info in row
        """

        time_list = []

        for i in range(2, self.data.max_row):
            if self.data.cell(i, 2).value >= dtime:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i - 1, j).value)
                time_list.append(name_info)
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                time_list.append(name_info)
                break
        return time_list

    def get_eventdata_byplayer(self, name):
        """get event data by player name

        Args:
            name (string): player's name

        Returns:
            list[list]: list of info in row with corresponding player's name
        """

        name_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 7).value.find(name) != -1:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                name_list.append(name_info)
        return name_list

    def get_eventdata_byteam(self, teamname):
        """ get event data by searching team name

        Args:
            teamname (string): team name

        Returns:
            list[list]: list of info in row with corresponding team name
        """

        name_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 8).value.find(teamname) != -1:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                name_list.append(name_info)
        return name_list

    def get_position_byevent(self, event):
        """get position data by event

        Args:
            event (string): name of event

        Returns:
            list: list of position with corresponding event 
        """
        pos_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 1).value == event:
                pos_info = self.data.cell(i, 9).value
                pos_list.append(pos_info)
        return pos_list


class Get_Videodata:
    """Get_Videodata."""

    def __init__(self, video_path, time_offset=0):
        """initialization

        Args:
            video_path (_string_): path of video
            time_offset (_float_): secs, time of beginning of game in the video
        """
        super(Get_Videodata, self).__init__()
        self.video_path = video_path
        self.video = cv2.VideoCapture(self.video_path)
        self.fps = self.video.get(cv2.CAP_PROP_FPS)
        self.time_offset = time_offset * self.fps  #number of f rames
        self.num_frame = self.video.get(7)

    def get_videodata_bytime(self, steps_frame, time: datetime.time):
        """ get video data by time

        Args:
            steps_frame (int): number of frame around event
            time (datetime.time): time

        Returns:
            frame
        """
        #time in second
        time = time.hour * 60 * 60 + time.minute * 60 + time.second
        time = (time * self.fps - steps_frame, time * self.fps + steps_frame)
        t1 = int(time[0] + self.time_offset)
        # t2 = int(time[1] + self.time_offset)
        self.video.set(cv2.CAP_PROP_POS_FRAMES, int(t1))
        j, frame = self.video.read()
        # cv2.imshow('sd', frame)
        # cv2.waitKey()
        # while self.video.isOpened():
        #     j, frame = self.video.read()
        #     if j:
        #         i += 1
        #         if i >= t1 and i <= t2:
        #             out.append(frame)
        #         elif i > t2:
        #             return out
        #     else:
        #         print('errors occured')
        return frame


if __name__ == '__main__':
    pass
    # path = 'datasets/IK Frej Täby-IF Brommapojkarna(0-1).xlsx'
    # xlsx = Get_Eventdata(path)
    # name = 'Shots'
    # # dtime = datetime.time(0, 8, 29)
    # # name = 'Christer Gustafsson'
    # # name = 'IF Brommapojkarna'
    # out = xlsx.get_position_byevent(name)
    # print(out)
    # print('=============')
    # print(out[0])
    path = 'datasets/ettan_test.mp4'
    dtime = datetime.time(0, 0, 46)
    vi = Get_Videodata(path, 0)
    out = vi.get_videodata_bytime(2, dtime)
    print(len(out))
