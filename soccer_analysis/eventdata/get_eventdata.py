#!/usr/bin/env python
# -*- encoding: utf-8 -*-
'''
@File    :   get_eventdata.py
@Time    :   2023/03/14 15:07:49
@Author  :   Rongfei Pan
@Version :   1.0
@Contact :   rongfei@kth.se 1838863836prf@gmail.com
@Desc    :   function library for event data
'''

# here put the import lib

from openpyxl import load_workbook
import datetime
import cv2
import json

from loguru import logger


class Get_Eventdata:

    def __init__(self, path):
        """initialized by path

        Args:
            path (string): path of xlsx file
        """
        super(Get_Eventdata, self).__init__()
        self.path = path
        self.wb = load_workbook(path)
        self.data = self.wb['Sheet1']

    def get_teamname(self):
        team_list = []
        i = 0
        while True:
            d = self.getdata_byaxis(2 + i, 8)
            if i == 0:
                team_list.append(d)
            else:
                if team_list[0] == d:
                    pass
                else:
                    team_list.append(d)
                    return team_list
            i += 1

    def get_initdirection(self):
        init_team = self.getdata_byaxis(2, 8)
        field_pos = self.getdata_byaxis(2, 9)  #9,12
        receiver_pos = self.getdata_byaxis(2, 12)
        field_pos = field_pos.split(';')
        receiver_pos = receiver_pos.split(';')
        for i, pos in enumerate(field_pos):
            field_pos[i] = int(pos)
        for i, pos in enumerate(receiver_pos):
            receiver_pos[i] = int(pos)
        if receiver_pos[0] > field_pos[0]:
            return (init_team, 'left')
        else:
            return (init_team, 'right')

    def getdata_byaxis(self, x, y):
        d = self.data.cell(x, y)
        return d.value

    def get_eventdata_byevent(self, name):
        """get event data by event

        Args:
            name (string): type of event

        Returns:
            list[list]: list of info in row with corresponding name
        """
        name_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 1).value == name:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                name_list.append(name_info)
        return name_list

    def get_eventdata_bytime(self, dtime):
        """get event data by time

        Args:
            dtime (datetime.time): time, Exp: datetime.time(xx, xx, xx)

        Returns:
            list[list]: list with two sublist with closest time. Sublist: info in row
        """

        time_list = []

        for i in range(2, self.data.max_row):
            if self.data.cell(i, 2).value >= dtime:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i - 1, j).value)
                time_list.append(name_info)
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                time_list.append(name_info)
                break
        return time_list

    def get_eventdata_byplayer(self, name):
        """get event data by player name

        Args:
            name (string): player's name

        Returns:
            list[list]: list of info in row with corresponding player's name
        """

        name_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 7).value.find(name) != -1:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                name_list.append(name_info)
        return name_list

    def get_eventdata_byteam(self, teamname):
        """ get event data by searching team name

        Args:
            teamname (string): team name

        Returns:
            list[list]: list of info in row with corresponding team name
        """

        name_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 8).value.find(teamname) != -1:
                name_info = []
                for j in range(1, self.data.max_column):
                    name_info.append(self.data.cell(i, j).value)
                name_list.append(name_info)
        return name_list

    def get_position_byevent(self, event):
        """get position data by event

        Args:
            event (string): name of event

        Returns:
            list: list of position with corresponding event 
        """
        pos_list = []
        for i in range(2, self.data.max_row):
            if self.data.cell(i, 1).value == event:
                pos_info = self.data.cell(i, 9).value
                pos_list.append(pos_info)
        return pos_list

    def get_startime(self):
        starttime = self.getdata_byaxis(2, 2)
        return starttime

    def get_direction_byteam(self, time, team_name):
        s = self.get_initdirection()
        t = self.get_startime()
        if datetime.datetime.combine(
                datetime.date.min, time) - datetime.datetime.combine(
                    datetime.date.min, t) >= datetime.timedelta(minutes=53):
            #second half
            if team_name == s[0]:
                if s[1] == 'left':
                    return 'right'
                else:
                    return 'left'
            else:
                return s[1]
        else:
            #first half
            if team_name == s[0]:
                return s[1]
            else:
                if s[1] == 'left':
                    return 'right'
                else:
                    return 'left'


class Get_Videodata:
    """Get_Videodata."""

    def __init__(self, video_path, time_offset=0):
        """initialization

        Args:
            video_path (_string_): path of video
            time_offset (_float_): secs, time of beginning of game in the video
        """
        super(Get_Videodata, self).__init__()
        self.video_path = video_path
        self.video = cv2.VideoCapture(self.video_path)
        self.fps = self.video.get(cv2.CAP_PROP_FPS)
        self.time_offset = time_offset * self.fps  #number of f rames
        self.num_frame = self.video.get(7)

    def get_videodata_bytime(self, steps_frame, time: datetime.time):
        """ get video data by time

        Args:
            steps_frame (int): number of frame around event
            time (datetime.time): time

        Returns:
            frame (list): single image
        """
        #time in second
        time = time.hour * 60 * 60 + time.minute * 60 + time.second
        time = (time * self.fps - steps_frame, time * self.fps + steps_frame)
        t1 = int(time[0] + self.time_offset)
        # t2 = int(time[1] + self.time_offset)
        self.video.set(cv2.CAP_PROP_POS_FRAMES, int(t1))
        j, frame = self.video.read()
        # cv2.imshow('sd', frame)
        # cv2.waitKey()
        # while self.video.isOpened():
        #     j, frame = self.video.read()
        #     if j:
        #         i += 1
        #         if i >= t1 and i <= t2:
        #             out.append(frame)
        #         elif i > t2:
        #             return out
        #     else:
        #         print('errors occured')
        return frame


class Get_xGdata(object):
    """docstring for Get_xGdata in json file."""

    def __init__(self, filename=''):
        """init get_xGdata

        Args:
            filename (str, optional): json file name with match data. Defaults to ''.
        """
        super(Get_xGdata, self).__init__()
        with open(filename) as f:
            self.data = json.load(f)

    def get_jsondata_shot(self):
        """get shot data in json file

        Returns:
            list: list of dict with shot event data
        """

        event_list = self.data['events']
        shot_list = []
        for i, event in enumerate(event_list):
            ac = event['action']
            if ac == 'Shot':
                shot_list.append(event)
        return shot_list


if __name__ == '__main__':
    path = 'datasets/test2/Assyriska FF - Täby FK (2-7).xlsx'
    xlsx = Get_Eventdata(path)
    name = 'Shots'
    # dtime = datetime.time(0, 8, 29)
    # name = 'Christer Gustafsson'
    # name = 'IF Brommapojkarna'
    out = xlsx.get_startime()
    print(out, type(out))
    print('=============')
    # print(out[0])
    # path = 'datasets/ettan_test.mp4'
    # dtime = datetime.time(0, 0, 46)
    # vi = Get_Videodata(path, 0)
    # out = vi.get_videodata_bytime(2, dtime)

    # jsonfile = 'datasets/test2/ettan, 2021786157109941299-AssyriskaFF-TäbyFK.json'
    # xgdata = Get_xGdata(jsonfile)
    # shot_list = xgdata.get_jsondata_shot()
    # logger.info(shot_list[0])
